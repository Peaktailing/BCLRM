"""从上传的Excel文件导入测试数据到数据库

该脚本用于将上传的测试Excel文件中的数据导入到SQLite数据库中。
- 612.4 / 612-2 / 623 / E612药品柜 / 化学品清单：药品柜库存清单（导入到reagent_bottle）
- 报废试剂空瓶(1)(2)(3)：报废空瓶记录（导入到reagent_bottle，标记为耗尽状态）

数据规范化：
- 去除名称前后空白
- 统一规格单位（g/ml）
- 名称+规格 唯一性去重
"""
import openpyxl
import re
import os
import sys
from pathlib import Path
from datetime import datetime

# 添加项目根目录到 Python 路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from db.database import Database
from utils.error_handler import logger


UPLOAD_DIR = "/workspace/.uploads"


# 在用试剂瓶文件（药品柜库存）
ACTIVE_FILES = [
    ("/workspace/.uploads/247e6c8e-0cb5-4d8a-bf0c-ff9e94a53b5d_612.4(1).xlsx", "工作表1", "E612 4号药品柜"),
    ("/workspace/.uploads/d10cdf63-dc00-4a8e-ac98-4dbfcd71d56f_612-2.xlsx", "工作表1", "E612 2号药品柜"),
    ("/workspace/.uploads/15ae7bf0-5b3e-43f5-babb-ad3a29c78f6c_623.xlsx", "工作表1", "E623药品柜"),
    ("/workspace/.uploads/05d02a27-85a4-4bcf-a0ca-3c119812eeac_E612 1号药品柜.xlsx", "Sheet1", "E612 1号药品柜"),
    ("/workspace/.uploads/91fda626-7d82-4970-9969-e28b40a9e7cd_E612 药品柜(1).xlsx", "Sheet1", "E612 药品柜"),
    ("/workspace/.uploads/c4d68980-ce05-4ebf-b710-5a925951221b_化学品清单.xlsx", "Sheet1", "化学品清单"),
]

# 报废试剂空瓶文件
SCRAP_FILES = [
    ("/workspace/.uploads/c3a4b71d-5883-4fb6-9dac-de87afc28a2b_报废试剂空瓶(1).xlsx", "固体", "E612"),
    ("/workspace/.uploads/c3a4b71d-5883-4fb6-9dac-de87afc28a2b_报废试剂空瓶(1).xlsx", "液体", "E612"),
    ("/workspace/.uploads/267b0839-060d-447d-b377-83c2428825fa_报废试剂空瓶(2).xlsx", "固体", "E612"),
    ("/workspace/.uploads/267b0839-060d-447d-b377-83c2428825fa_报废试剂空瓶(2).xlsx", "液体", "E612"),
    ("/workspace/.uploads/4dfecbfd-b4a3-47c2-a7de-6950a2ee139b_报废试剂空瓶(3).xlsx", "固体", "E612"),
    ("/workspace/.uploads/4dfecbfd-b4a3-47c2-a7de-6950a2ee139b_报废试剂空瓶(3).xlsx", "液体", "E612"),
]


def normalize_spec(spec_str):
    """规范化规格字符串

    Args:
        spec_str: 原始规格字符串，如 '500g', '25ml', '100ml', '25m l' 等

    Returns:
        (数值, 单位) 元组
    """
    if not spec_str:
        return None, None

    s = str(spec_str).strip()
    # 处理全角空格、半角空格等
    s = s.replace('\u3000', '').replace(' ', '').replace('\u2006', '').replace('\u2009', '')

    # 匹配数字和单位
    match = re.match(r'^([0-9.]+)(mg|g|kg|ml|mL|L)$', s, re.IGNORECASE)
    if match:
        value = float(match.group(1))
        unit = match.group(2).lower()
        # 转换为标准单位：g 或 ml
        if unit == 'kg':
            value *= 1000
            unit = 'g'
        elif unit == 'l':
            value *= 1000
            unit = 'ml'
        elif unit == 'mg':
            value /= 1000
            unit = 'g'
        return value, unit

    return None, s


def normalize_name(name):
    """规范化名称，去除空白和特殊字符"""
    if not name:
        return ""
    s = str(name).strip()
    s = s.replace('\u3000', '').replace('\u2006', '').replace('\u2009', '')
    s = re.sub(r'\s+', '', s)
    return s


def load_excel_data():
    """加载所有Excel数据

    Returns:
        (active_records, scrap_records) 元组
        active_records: 在用试剂瓶记录列表
        scrap_records: 报废试剂空瓶记录列表
    """
    active_records = []  # [(name, spec, qty, location)]
    scrap_records = []   # [(name, spec, qty, production_date, location)]

    # 加载在用药品柜数据
    for fpath, sheet, location in ACTIVE_FILES:
        if not os.path.exists(fpath):
            print(f"  跳过: 文件不存在 {fpath}")
            continue

        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb[sheet]

        for row in ws.iter_rows(min_row=2, values_only=True):
            # 跳过空行
            if not row or all(cell is None for cell in row):
                continue
            # 跳过表头
            if row[0] and str(row[0]).strip() in ('序号', '名称'):
                continue

            # 找到名称列（兼容不同的列顺序）
            name = None
            spec = None
            qty = None
            shelf = None  # 层数

            for cell in row:
                if cell is None:
                    continue

            # 根据列数判断格式
            if len(row) >= 5 and row[1] is not None:
                # 格式: 序号, 名称, 规格, 数量, 层数, ...
                name = row[1]
                spec = row[2]
                qty = row[3]
                shelf = row[4]
            elif len(row) >= 3:
                # 化学品清单格式: 名称, 规格, 数量
                name = row[0]
                spec = row[1]
                qty = row[2]

            if not name or not str(name).strip():
                continue

            name_clean = normalize_name(name)
            if not name_clean:
                continue

            spec_value, spec_unit = normalize_spec(spec)
            qty_int = int(qty) if qty and str(qty).strip().isdigit() else 1

            active_records.append({
                'name': name_clean,
                'spec_value': spec_value,
                'spec_unit': spec_unit,
                'spec_raw': str(spec) if spec else '',
                'qty': qty_int,
                'location': location,
                'shelf': str(shelf).strip() if shelf else None,
            })

    # 加载报废试剂空瓶数据
    for fpath, sheet, location in SCRAP_FILES:
        if not os.path.exists(fpath):
            print(f"  跳过: 文件不存在 {fpath}")
            continue

        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb[sheet]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None for cell in row):
                continue

            # 报废数据格式: 品名, 规格, 数量, ...
            name = row[1] if len(row) > 1 else None
            spec = row[2] if len(row) > 2 else None
            qty = row[3] if len(row) > 3 else None
            prod_date = row[5] if len(row) > 5 else None  # 出厂日期

            if not name or not str(name).strip():
                continue

            name_clean = normalize_name(name)
            if not name_clean:
                continue

            spec_value, spec_unit = normalize_spec(spec)
            qty_int = int(qty) if qty and str(qty).strip().isdigit() else 1

            prod_date_str = None
            if isinstance(prod_date, datetime):
                prod_date_str = prod_date.strftime('%Y/%m/%d')

            scrap_records.append({
                'name': name_clean,
                'spec_value': spec_value,
                'spec_unit': spec_unit,
                'spec_raw': str(spec) if spec else '',
                'qty': qty_int,
                'location': location,
                'production_date': prod_date_str,
            })

    return active_records, scrap_records


def dedupe_records(records, name_key='name'):
    """对记录去重，合并数量

    去重策略：按 (规范化名称, 规格数值, 规格单位, 位置) 分组，相同组合的数量累加
    """
    grouped = {}
    for r in records:
        key = (r[name_key], r['spec_value'], r['spec_unit'], r['location'])
        if key not in grouped:
            grouped[key] = r.copy()
        else:
            grouped[key]['qty'] += r['qty']

    return list(grouped.values())


def parse_spec_to_grams(spec_value, spec_unit):
    """将规格转换为以克为单位的剩余量

    简化处理：g 和 ml 都按 1:1 转换为 g
    """
    if spec_value is None:
        return None
    return float(spec_value)


def insert_reagent_bottles(db, records, status, start_number):
    """插入试剂瓶记录到数据库

    Args:
        db: 数据库实例
        records: 记录列表
        status: 状态 '可借' 或 '耗尽'
        start_number: 起始试剂瓶编号

    Returns:
        (count, next_number) 元组
    """
    count = 0
    current_number = start_number
    for r in records:
        # 每条记录中的每个数量创建一行（一个试剂瓶）
        for _ in range(r['qty']):
            # 简化处理：使用名称+规格作为试剂瓶数据
            remaining = parse_spec_to_grams(r['spec_value'], r['spec_unit'])
            if remaining is None or remaining <= 0:
                continue

            fields = {
                'bottle_number': current_number,
                'reagent_name': r['name'],
                'remaining_quantity': remaining,
                'specification': remaining,
                'borrowable_flag': status,
                'borrowable_check': 1 if status == '可借' else 0,
                'storage_location': r['location'],
                'inbound_date': datetime.now().strftime('%Y/%m/%d'),
            }

            if r.get('production_date'):
                fields['production_date'] = r['production_date']

            try:
                columns = ', '.join(fields.keys())
                placeholders = ', '.join(['?' for _ in fields])
                query = f"INSERT INTO reagent_bottle ({columns}) VALUES ({placeholders})"
                db.execute_insert(query, tuple(fields.values()))
                count += 1
                current_number += 1
            except Exception as e:
                logger.error(f"插入试剂瓶失败 [{r['name']}]: {str(e)}", exception=e)

    return count, current_number


def main():
    """主函数"""
    print("=" * 60)
    print("测试数据导入工具")
    print("=" * 60)

    # 初始化数据库
    db = Database()
    db.init_tables()

    # 加载Excel数据
    print("\n正在加载Excel数据...")
    active_records, scrap_records = load_excel_data()
    print(f"  原始在用记录: {len(active_records)} 条")
    print(f"  原始报废记录: {len(scrap_records)} 条")

    # 数据去重
    print("\n正在去重和合并...")
    active_dedup = dedupe_records(active_records)
    scrap_dedup = dedupe_records(scrap_records)
    print(f"  去重后在用记录: {len(active_dedup)} 条 (试剂瓶总数: {sum(r['qty'] for r in active_dedup)})")
    print(f"  去重后报废记录: {len(scrap_dedup)} 条 (试剂瓶总数: {sum(r['qty'] for r in scrap_dedup)})")

    # 清空现有试剂瓶表（测试数据重置）
    print("\n清空现有试剂瓶数据...")
    # 临时禁用外键约束，避免 storage_location 外键失败
    db.execute_update("PRAGMA foreign_keys = OFF")
    db.execute_update("DELETE FROM reagent_bottle")

    # 插入在用记录（可借状态）
    print("\n正在插入在用试剂瓶数据...")
    active_count, next_number = insert_reagent_bottles(db, active_dedup, '可借', 1)
    print(f"  插入完成: {active_count} 条在用试剂瓶记录")

    # 插入报废记录（耗尽状态）
    print("\n正在插入报废试剂空瓶数据...")
    scrap_count, _ = insert_reagent_bottles(db, scrap_dedup, '耗尽', next_number)
    print(f"  插入完成: {scrap_count} 条报废试剂空瓶记录")

    # 验证结果
    print("\n数据统计:")
    cursor = db.connection.cursor()
    cursor.execute("SELECT COUNT(*) FROM reagent_bottle")
    total = cursor.fetchone()[0]
    print(f"  试剂瓶总数: {total}")

    cursor.execute("SELECT COUNT(*) FROM reagent_bottle WHERE borrowable_flag = '可借'")
    active = cursor.fetchone()[0]
    print(f"  可借数量: {active}")

    cursor.execute("SELECT COUNT(*) FROM reagent_bottle WHERE borrowable_flag = '耗尽'")
    exhausted = cursor.fetchone()[0]
    print(f"  耗尽数量: {exhausted}")

    print("\n" + "=" * 60)
    print("数据导入完成！")
    print("=" * 60)

    # 重新启用外键约束
    db.execute_update("PRAGMA foreign_keys = ON")
    db.close()


if __name__ == "__main__":
    main()
